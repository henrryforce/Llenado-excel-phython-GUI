# -*- coding: utf-8 -*-
"""
Spyder Editor

This is a temporary script file.
"""

from tkinter import *
import openpyxl
a=[]
cont=[]
nRow=0
nameArch= ''
numID=0
#Recolecta la informacion del Input y lo convierte en una lista tambien vacia el input
def conver():
    global a
    global cont
    global nRow
    nRow+=1
    b=texto.get(1.0, 'end')
    a= b.replace('\n',',')
    a=a.split(sep=',')
    a.pop(len(a)-1)
    print(a)
    cont.append(a)
    texto.delete(1.0, 'end')
    
    '''
    book = openpyxl.load_workbook('ejercicios528.xlsx',data_only=False)
    hoja = book.active
    for i in range(0,10):
        hoja.cell(row = 3,column =4+i, value=int(a[i]))
    book.save('ejercicios529.xlsx')'''
def agrega():
    global a
    global cont
    global nRow
    global nameArch,numID
    print(cont)
    print(nRow)
    print(entry.get())
    nameArch=entry.get()
    numID=entry1.get()
    book = openpyxl.load_workbook('ejercicios528.xlsx',data_only=False)
    hoja = book.active
    for j in range(0,nRow):        
        for i in range(0,10):
            hoja.cell(row = 2+j,column =4+i, value=int(cont[j][i]))
    for j in range(0,nRow):
        hoja.cell(row=2+j,column=1,value=int(numID))
    book.save(nameArch+numID+'.xlsx')
    entry.delete(0, 'end')
    entry1.delete(0, 'end')
    a=[]
    cont=[]
    nRow=0
    nameArch= ''
    numID=0
root = Tk()
root.title("Llena Excel 1.5.0") 
label=Label(root, text="Llenar excel con Python")
label.pack(anchor=CENTER)
label.config(fg="blue",   # Background
             font=("Verdana",24))
texto = Text(root)
texto.pack()
texto.config(width=30, height=10, font=("Consolas",12),padx=15, pady=15, selectbackground="red")

Button(root, text="Agregar", command=conver).pack()
Button(root, text="Agrgar a excel", command=agrega,).pack(side=BOTTOM)
entry = Entry(root,textvariable=nameArch)
entry.pack(side=RIGHT)
label = Label(root, text="Nombre del archivo")
label.pack(side=RIGHT)
entry1 = Entry(root,textvariable=numID)
entry1.pack(side=RIGHT)
label1 = Label(root, text="Numero de ID")
label1.pack(side=LEFT)




root.mainloop()
